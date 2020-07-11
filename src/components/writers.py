from openpyxl import Workbook

COLUMNS = dict(current=['Field', 'Value'],
               previous=['Field', 'Value'],
               difference=['Field', 'Type', 'Current', 'Prevision'])

TITLES = ['current', 'previous', 'difference']

class xlsxWriter(Workbook):
    def __init__(self, query):
        self.workbook = Workbook()
        self.current = query.result
        self.previous = query.previous
        self.difference = query.difference

    def do(self):
        for title in TITLES:
            self.do_worksheet(title)

    def save(self, output):
        self.workbook.save(output)


    def do_worksheet(self, title):

        if 'current' in title and not self.current: return
        if 'previous' in title and not self.previous: return
        if 'difference' in title and not self.difference: return

        worksheet = self.workbook.active if 'current' in title else self.workbook.create_sheet(title)
        worksheet.title = title.capitalize()

        columns = COLUMNS.get(title)
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

            # Iterate through data

        if 'difference' in title:
            rows = [[d.get('field'), d.get('type'), d.get('current'), d.get('previous')] for d in self.difference]

        else:
            data = self.current if 'current' in title else self.previous
            rows = [[field, value]  for field,value in data.items()]

        for row in rows:
            print(rows)
            row_num += 1
            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value